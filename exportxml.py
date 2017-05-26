#!/usr/bin/env python
#  encoding=utf-8
#  pip install openpyxl

import os, os.path
import sys
import re
import argparse
import json
import shutil
from openpyxl import load_workbook
from function import readfile, writefile, copyfile

# <com.g2d.studio.ui.edit.gui.UELabel Attributes="" ImageFont="" clip_local_bounds="false" clipbounds="false" enable="true" enable_childs="true" height="30" local_bounds="0,0,60,30" lock="true" name="lb_title" text="动作" textBorderAlpha="0.0" textBorderColor="0" textColor="ffffffff" textFont="" textFontSize="20" text_anchor="C_C" text_offset_x="0" text_offset_y="0" uiAnchor="" uiEffect="" userData="" userTag="0" visible="true" visible_content="true" width="60" x="130.0" y="4.0" z="0.0"/>
# <com.g2d.studio.ui.edit.gui.UEButton Attributes="" ImageFont="" clip_local_bounds="false" clipbounds="false" enable="true" enable_childs="true" focusTextColor="ffffffff" height="35" imageAnchor="C_C" imageAtlasDown="" imageAtlasUp="" imageOffsetX="0" imageOffsetY="0" imageTextDown="" imageTextUp="" local_bounds="0,0,35,35" lock="true" name="btn_close" text="" textBorderAlpha="100.0" textBorderColor="ff000000" textDown="" textFont="" textSize="0" text_anchor="C_C" text_offset_x="0" text_offset_y="0" uiAnchor="" uiEffect="" unfocusTextColor="ffffffff" userData="" userTag="0" visible="true" visible_content="true" width="35" x="807.0" y="-3.0" z="0.0">
# <com.g2d.studio.ui.edit.gui.UEToggleButton Attributes="" ImageFont="" clip_local_bounds="false" clipbounds="false" enable="true" enable_childs="true" focusTextColor="ffffffff" height="78" imageAnchor="C_C" imageAtlasDown="#dynamic/associate/output/associate.xml|associate|9" imageAtlasUp="#dynamic/associate/output/associate.xml|associate|11" imageOffsetX="-15" imageOffsetY="0" imageTextDown="" imageTextUp="" isChecked="true" local_bounds="0,0,121,78" lock="true" name="tbt_an1" text="" textBorderAlpha="100.0" textBorderColor="ff000000" textDown="" textFont="" textSize="0" text_anchor="C_C" text_offset_x="0" text_offset_y="0" uiAnchor="" uiEffect="" unfocusTextColor="ffffffff" userData="" userTag="0" visible="true" visible_content="true" width="121" x="0.0" y="30.0" z="0.0">
P_LABEL = r'^(<com\.g2d\.studio\.ui\.edit\.gui\.(UELabel|UEButton|UEToggleButton) .*? name=")([^"]+)(" .*text=")([^"]+)(".*)$'
# P_BUTTON = r'^(<com\.g2d\.studio\.ui\.edit\.gui\.UEButton .*? name=")(btn_([^"]+))(" .*text=")([^"]+)(".*)$'
# P_TBUTTON = r'^(<com\.g2d\.studio\.ui\.edit\.gui\.UEToggleButton .*? name=")(tbt_([^"]+))(" .*text=")([^"]+)(".*)$'


errors = []

class Config():
    __slots__ = ("excelDir", "xmlDir", "backXmlDir", "excelTemplate", "out")
    def __init__(self):
        pass

    def load(self, path):
        jsonObj = json.loads(readfile(path))
        self.excelDir = jsonObj['excelDir'].decode('utf-8').encode('gb2312')
        self.xmlDir = jsonObj['xmlDir'].decode('utf-8').encode('gb2312')
        self.backXmlDir = jsonObj['backXmlDir'].decode('utf-8').encode('gb2312')
        self.excelTemplate = jsonObj['excelTemplate'].decode('utf-8').encode('gb2312')
        self.out = jsonObj['out'].decode('utf-8').encode('gb2312')

    def parse(self, args):
        for k in self.__slots__:
            print(k)
            if args[k] is not None:
                self[k] = args[k].decode('utf-8').encode('gb2312')

def doKeys(xmlFile, duplicateKeys):
    global errors
    errors.append((xmlFile, duplicateKeys.items()))

def replaceXml(data, prefix):
    f = lambda g:"".join([g.group(1),g.group(3),g.group(4),prefix,g.group(2),g.group(6)])
    return re.sub(P_LABEL, f, data, 0, re.M)

def checkDuplicateName(name, map, errMap):
    count = map.get(name, 0) + 1
    map[name] = count
    if count > 1:
        errMap[name] = count
    return count > 1

def guixml2excel(xmlFile, backXmlFile, excelFile, prefix):
    data = readfile(xmlFile)
    matches = re.findall(P_LABEL, data, re.M)
    if len(matches) == 0:
        return

    writefile(backXmlFile, replaceXml(data, prefix))
    shutil.copyfile(config.excelTemplate, excelFile)

    
    # print(str(matches).encode('utf-8'))
    keys = dict()

    # <(name,value)>
    pairs = []
    duplicateKeys = dict()
    for match in matches:
        type = match[1]
        name = match[2]
        if checkDuplicateName(name, keys, duplicateKeys):
            continue
        pairs.append((name, match[4]))
    pairs.sort(key = lambda p: p[0])

    row = 2
    wb = load_workbook(excelFile, keep_vba = True)
    ws = wb.active
    for p in pairs:
        c = ws.cell(row=row, column=1)
        c.value = prefix + p[0]
        c = ws.cell(row=row, column=2)
        c.value = p[1]
        row = row + 1
        # print(match[0], match[1], match[2], type(match[2]))
    print(prefix, row - 2)
    wb.save(excelFile)
    if len(duplicateKeys) > 0:
        doKeys(xmlFile, duplicateKeys)

def doTask(root, file, config):
    name = file[0:-8]
    prefix = name + "_"
    xmlFile = os.path.join(root, file)
    excelFile = os.path.join(config.excelDir, '%s.xlsm' % name)
    relpath = os.path.relpath(xmlFile, config.xmlDir)
    backXmlFile = os.path.join(config.backXmlDir, relpath)
    copyfile(xmlFile, backXmlFile)
    guixml2excel(xmlFile, backXmlFile, excelFile, prefix)

# -------------- main ----------------
if __name__ == '__main__':
    parser = argparse.ArgumentParser(usage='%(prog)s [options] config',
        description='export xmlui to excel')
    parser.add_argument('config', nargs = '?',
        help='config file')
    parser.add_argument('--excelDir', 
        help='excelDir')
    parser.add_argument('--xmlDir', 
        help='xmlDir')
    parser.add_argument('--backXmlDir', 
        help='backXmlDir')
    parser.add_argument('--excelTemplate', 
        help='excelTemplate')
    parser.add_argument('-o', '--out', 
        help='out')

    # ("excelDir", "xmlDir", "backXmlDir", "excelTemplate", "log")
    args = parser.parse_args()
    # print(args.config)
    # 
    config = Config()
    if args.config is not None:
        configPath = os.path.abspath(args.config)
        if not configPath.endswith('.json'):
            configPath = configPath + '.json'
        config.load(configPath)

    config.parse(vars(args))

    # global errors
    errors = []

    if not os.path.exists(config.excelDir):
        os.mkdir(config.excelDir)

    if not os.path.exists(config.backXmlDir):
        os.mkdir(config.backXmlDir)

    if os.path.exists(config.out):
        os.remove(config.out)

    for root, dirs, files in os.walk(config.xmlDir):
        for f in files:
            if f.endswith('.gui.xml'):
                doTask(root, f, config)

    if len(errors) > 0:
        from renderhtml import renderhtml
        renderhtml(errors, ("key", "count"), config.out)

